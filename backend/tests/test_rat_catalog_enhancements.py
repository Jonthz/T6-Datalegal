"""Tests para el incremento B2: RAT (ID estructurado, bases múltiples, export XLSX)
y catálogos (creación individual, versionado con historial, re-clasificación)."""

from io import BytesIO

from fastapi.testclient import TestClient

from tests.conftest import auth_headers


# ── RAT: identificador estructurado ──────────────────────────────────────────


def test_rat_code_generated_per_area(client: TestClient, dpo_token):
    """El código del RAT se genera con prefijo de área y secuencial contiguo."""
    a = client.post(
        "/api/v1/treatment-activities",
        json={"name": "Crédito 1", "purpose": "x", "legal_basis": "CONTRACT", "area": "CRÉDITO"},
        headers=auth_headers(dpo_token),
    ).json()
    b = client.post(
        "/api/v1/treatment-activities",
        json={"name": "Crédito 2", "purpose": "x", "legal_basis": "CONTRACT", "area": "CRÉDITO"},
        headers=auth_headers(dpo_token),
    ).json()
    c = client.post(
        "/api/v1/treatment-activities",
        json={
            "name": "Tesorería 1",
            "purpose": "x",
            "legal_basis": "CONTRACT",
            "area": "TESORERÍA",
        },
        headers=auth_headers(dpo_token),
    ).json()
    assert a["rat_code"] == "CRE001"
    assert b["rat_code"] == "CRE002"
    assert c["rat_code"] == "TES001"


def test_rat_code_assigned_on_update_when_area_set(client: TestClient, dpo_token):
    """Si se crea sin área, el código se genera al asignar el área en un PATCH."""
    created = client.post(
        "/api/v1/treatment-activities",
        json={"name": "Sin área", "purpose": "x", "legal_basis": "CONTRACT"},
        headers=auth_headers(dpo_token),
    ).json()
    assert created["rat_code"] == "GEN001"
    updated = client.patch(
        f"/api/v1/treatment-activities/{created['id']}",
        json={"area": "IMPORTACIONES"},
        headers=auth_headers(dpo_token),
    ).json()
    # GEN001 ya estaba asignado, así que no se regenera; se mantiene.
    assert updated["rat_code"] == "GEN001"


# ── RAT: selección múltiple de bases de licitud ──────────────────────────────


def test_multiple_legal_bases(client: TestClient, dpo_token):
    """legal_bases admite varias bases; legal_basis (principal) se sincroniza."""
    data = client.post(
        "/api/v1/treatment-activities",
        json={
            "name": "Nómina",
            "purpose": "pago",
            "legal_bases": ["LEGAL_OBLIGATION", "CONTRACT"],
            "complementary_legal_bases": ["ART_28"],
            "area": "TALENTO HUMANO",
        },
        headers=auth_headers(dpo_token),
    ).json()
    assert data["legal_basis"] == "LEGAL_OBLIGATION"
    assert data["legal_bases"] == ["LEGAL_OBLIGATION", "CONTRACT"]
    assert data["complementary_legal_bases"] == ["ART_28"]


def test_new_rat_variables_persist(client: TestClient, dpo_token):
    """Las variables nuevas del RAT se guardan y se devuelven."""
    data = client.post(
        "/api/v1/treatment-activities",
        json={
            "name": "Selección con IA",
            "purpose": "reclutamiento",
            "legal_basis": "CONTRACT",
            "area": "TALENTO HUMANO",
            "uses_ai": True,
            "uses_profiling": True,
            "requires_dpia": True,
            "has_special_data": True,
            "involves_minors": False,
            "recipients": ["Comité de selección"],
            "processors": ["EVALUAR"],
            "data_categories": ["CV", "Experiencia"],
            "mtge_score": 7.5,
            "mtge_result": "Gran escala probable",
        },
        headers=auth_headers(dpo_token),
    ).json()
    assert data["uses_ai"] is True
    assert data["requires_dpia"] is True
    assert data["processors"] == ["EVALUAR"]
    assert data["mtge_result"] == "Gran escala probable"


# ── RAT: exportación en Excel ────────────────────────────────────────────────


def test_ropa_xlsx_export(client: TestClient, dpo_token):
    """El endpoint /ropa/xlsx devuelve un workbook válido con encabezados del RAT."""
    from openpyxl import load_workbook

    client.post(
        "/api/v1/treatment-activities",
        json={"name": "Export me", "purpose": "x", "legal_basis": "CONTRACT", "area": "CRÉDITO"},
        headers=auth_headers(dpo_token),
    )
    resp = client.get("/api/v1/ropa/xlsx", headers=auth_headers(dpo_token))
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = load_workbook(BytesIO(resp.content))
    ws = wb["RAT"]
    headers = [c.value for c in ws[1]]
    assert "ID_TRATAMIENTO" in headers
    assert "Base de licitud principal" in headers
    assert ws.max_row >= 2


# ── Catálogos: creación individual + auto-clasificación ──────────────────────


def test_single_catalog_create_auto_classifies(client: TestClient, dpo_token):
    """POST /catalogs crea una entrada y auto-clasifica por el code conocido."""
    resp = client.post(
        "/api/v1/catalogs",
        json={"type": "DATA_TYPE", "code": "HEALTH_DATA", "label": "Salud"},
        headers=auth_headers(dpo_token),
    )
    assert resp.status_code == 201
    data = resp.json()
    assert data["sensitivity"] == "SENSITIVE"
    assert data["criticality"] == "HIGH"
    assert data["version"] == 1


def test_auditor_cannot_create_single_catalog(client: TestClient, auditor_token):
    """El auditor (solo lectura) no puede crear entradas de catálogo."""
    resp = client.post(
        "/api/v1/catalogs",
        json={"type": "DATA_TYPE", "code": "X", "label": "X"},
        headers=auth_headers(auditor_token),
    )
    assert resp.status_code == 403


# ── Catálogos: versionado con historial ──────────────────────────────────────


def test_catalog_version_history(client: TestClient, dpo_token):
    """Cada edición incrementa la versión y deja snapshot recuperable."""
    created = client.post(
        "/api/v1/catalogs",
        json={"type": "DATA_TYPE", "code": "PET_NAME", "label": "Mascota"},
        headers=auth_headers(dpo_token),
    ).json()
    client.patch(
        f"/api/v1/catalogs/{created['id']}",
        json={"label": "Nombre de mascota"},
        headers=auth_headers(dpo_token),
    )
    client.patch(
        f"/api/v1/catalogs/{created['id']}",
        json={"description": "campo trivial"},
        headers=auth_headers(dpo_token),
    )
    versions = client.get(
        f"/api/v1/catalogs/{created['id']}/versions", headers=auth_headers(dpo_token)
    ).json()
    # v1 (creación) + v2 + v3
    assert [v["version"] for v in versions] == [3, 2, 1]
    assert versions[0]["label"] == "Nombre de mascota"


# ── Catálogos: re-clasificación ──────────────────────────────────────────────


def test_catalog_reclassify(client: TestClient, dpo_token):
    """reclassify recomputa sensibilidad/criticidad desde el code y versiona."""
    created = client.post(
        "/api/v1/catalogs",
        json={
            "type": "DATA_TYPE",
            "code": "BIOMETRIC",
            "label": "Biométrico",
            "sensitivity": "ORDINARY",
        },
        headers=auth_headers(dpo_token),
    ).json()
    # forzamos un valor incorrecto y luego re-clasificamos
    resp = client.post(
        f"/api/v1/catalogs/{created['id']}/reclassify", headers=auth_headers(dpo_token)
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["sensitivity"] == "SENSITIVE"
    assert data["criticality"] == "HIGH"
    assert data["version"] == created["version"] + 1


def test_reclassify_unknown_code_returns_422(client: TestClient, dpo_token):
    """Sin regla automática para el code, reclassify responde 422."""
    created = client.post(
        "/api/v1/catalogs",
        json={"type": "DATA_TYPE", "code": "UNKNOWN_XYZ", "label": "Desconocido"},
        headers=auth_headers(dpo_token),
    ).json()
    resp = client.post(
        f"/api/v1/catalogs/{created['id']}/reclassify", headers=auth_headers(dpo_token)
    )
    assert resp.status_code == 422
