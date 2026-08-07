"""US-RF31-1: ROPA report generation from registered treatment activities."""

from datetime import datetime, timezone
from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from fpdf import FPDF
from sqlalchemy.orm import Session

from app.api.deps import get_current_tenant_id, get_db, require_permission
from app.models.treatment_activity import TreatmentActivity
from app.models.user import User

router = APIRouter(prefix="/ropa", tags=["ropa"])


@router.get("")
def get_ropa(
    _: Annotated[User, Depends(require_permission("ropa", "r"))],
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
) -> dict:
    """Return the Record of Processing Activities (ROPA) report as JSON (US-RF31-1)."""
    return _build_ropa_data(tenant_id, db)


@router.get("/pdf")
def get_ropa_pdf(
    _: Annotated[User, Depends(require_permission("ropa", "r"))],
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Return the ROPA report as a downloadable PDF (US-RF31-1)."""
    ropa_data = _build_ropa_data(tenant_id, db)
    pdf_bytes = _generate_ropa_pdf(ropa_data)
    return StreamingResponse(
        BytesIO(bytes(pdf_bytes)),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=ropa_report.pdf"},
    )


@router.get("/xlsx")
def get_ropa_xlsx(
    _: Annotated[User, Depends(require_permission("ropa", "r"))],
    tenant_id: int = Depends(get_current_tenant_id),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Return the RAT/ROPA as a downloadable Excel workbook (US-RF31-1)."""
    activities = (
        db.query(TreatmentActivity)
        .filter(TreatmentActivity.tenant_id == tenant_id)
        .order_by(TreatmentActivity.rat_code, TreatmentActivity.id)
        .all()
    )
    xlsx_bytes = _generate_ropa_xlsx(activities)
    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": "attachment; filename=RAT.xlsx"},
    )


# Columnas del RAT: (encabezado, atributo/callable). Alineadas con la plantilla LOPDP.
_RAT_COLUMNS: list[tuple[str, str]] = [
    ("ID_TRATAMIENTO", "rat_code"),
    ("Nombre del tratamiento", "name"),
    ("Área responsable", "area"),
    ("Responsable operativo", "operational_owner"),
    ("Finalidad específica", "purpose"),
    ("Categorías de titulares", "data_subjects"),
    ("Categorías de datos", "data_categories"),
    ("Origen de los datos", "data_origin"),
    ("Operaciones de tratamiento", "treatment_operations"),
    ("Perfilamiento", "uses_profiling"),
    ("Uso de IA", "uses_ai"),
    ("Decisión automatizada", "automated_decision"),
    ("Base de licitud principal", "legal_basis"),
    ("Bases de licitud (todas)", "legal_bases"),
    ("Bases complementarias/condicionadas", "complementary_legal_bases"),
    ("Encargado(s)", "processors"),
    ("Destinatario(s)", "recipients"),
    ("Transferencia internacional", "is_cross_border"),
    ("País / mecanismo", "destination_countries"),
    ("Plazo de conservación (días)", "retention_period_days"),
    ("Datos especiales", "has_special_data"),
    ("Datos de NNA", "involves_minors"),
    ("EIPD", "requires_dpia"),
    ("Sistema / plataforma", "system_platform"),
    ("Medidas técnicas", "technical_measures"),
    ("Medidas organizativas", "organizational_measures"),
    ("Medidas físicas", "physical_measures"),
    ("Medidas jurídicas", "legal_measures"),
    ("Puntaje MTGE", "mtge_score"),
    ("Resultado MTGE", "mtge_result"),
    ("Estado", "status"),
]


def _cell_value(activity: TreatmentActivity, attr: str):
    """Render a treatment-activity attribute into an Excel-friendly value."""
    value = getattr(activity, attr, None)
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return value


def _generate_ropa_xlsx(activities: list[TreatmentActivity]) -> bytes:
    """Render the RAT to an .xlsx workbook using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "RAT"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)

    headers = [h for h, _ in _RAT_COLUMNS]
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 26

    for activity in activities:
        ws.append([_cell_value(activity, attr) for _, attr in _RAT_COLUMNS])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_ropa_data(tenant_id: int, db: Session) -> dict:
    """Build the ROPA data structure grouped by legal basis."""
    activities = db.query(TreatmentActivity).filter(TreatmentActivity.tenant_id == tenant_id).all()
    activities_by_basis: dict[str, list] = {}
    for act in activities:
        entry = {
            "id": act.id,
            "name": act.name,
            "purpose": act.purpose,
            "legal_basis": act.legal_basis,
            "personal_data_types": act.personal_data_types,
            "data_subjects": act.data_subjects,
            "is_cross_border": act.is_cross_border,
            "destination_countries": act.destination_countries,
            "processor_name": act.processor_name,
            "processor_country": act.processor_country,
            "retention_period_days": act.retention_period_days,
            "status": act.status,
            "created_at": act.created_at.isoformat() if act.created_at else None,
        }
        activities_by_basis.setdefault(act.legal_basis, []).append(entry)

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "tenant_id": tenant_id,
        "total_activities": len(activities),
        "activities_by_legal_basis": activities_by_basis,
    }


def _generate_ropa_pdf(ropa_data: dict) -> bytes:
    """Render the ROPA data dict to PDF bytes using fpdf2."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(
        0,
        10,
        "Record of Processing Activities (ROPA)",
        new_x="LMARGIN",
        new_y="NEXT",
        align="C",
    )
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(
        0,
        8,
        f"Generated: {ropa_data['generated_at']}",
        new_x="LMARGIN",
        new_y="NEXT",
        align="C",
    )
    pdf.cell(
        0,
        8,
        f"Tenant ID: {ropa_data['tenant_id']}  Total activities: {ropa_data['total_activities']}",
        new_x="LMARGIN",
        new_y="NEXT",
        align="C",
    )
    pdf.ln(6)

    for legal_basis, entries in ropa_data["activities_by_legal_basis"].items():
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(
            0,
            9,
            f"Legal Basis: {legal_basis} ({len(entries)} activities)",
            new_x="LMARGIN",
            new_y="NEXT",
            fill=True,
        )
        pdf.ln(2)

        for act in entries:
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(
                0,
                7,
                f"[{act['id']}] {act['name']} - Status: {act['status']}",
                new_x="LMARGIN",
                new_y="NEXT",
            )
            pdf.set_font("Helvetica", "", 10)
            pdf.multi_cell(0, 6, f"Purpose: {act['purpose']}")
            pdf.cell(
                0,
                6,
                f"Retention: {act['retention_period_days']} days",
                new_x="LMARGIN",
                new_y="NEXT",
            )
            types_str = (
                ", ".join(act["personal_data_types"]) if act["personal_data_types"] else "N/A"
            )
            pdf.cell(0, 6, f"Data types: {types_str}", new_x="LMARGIN", new_y="NEXT")
            subjects_str = ", ".join(act["data_subjects"]) if act["data_subjects"] else "N/A"
            pdf.cell(0, 6, f"Data subjects: {subjects_str}", new_x="LMARGIN", new_y="NEXT")
            if act["is_cross_border"]:
                destinations = act["destination_countries"]
                countries = ", ".join(destinations) if destinations else "N/A"
                pdf.cell(
                    0,
                    6,
                    f"Cross-border - Destinations: {countries}",
                    new_x="LMARGIN",
                    new_y="NEXT",
                )
            if act["processor_name"]:
                pdf.cell(
                    0,
                    6,
                    f"Processor: {act['processor_name']} ({act['processor_country'] or 'N/A'})",
                    new_x="LMARGIN",
                    new_y="NEXT",
                )
            pdf.ln(3)
        pdf.ln(4)

    return pdf.output()
